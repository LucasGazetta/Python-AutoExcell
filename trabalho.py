from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "João": {
        "Matemática": 65,
        "Geografia": 78,
        "Inglês": 87,
        "História": 98
    },
    "Maria": {
        "Matemática": 54,
        "Geografia": 73,
        "Inglês": 82,
        "História": 65
    },
    "Lucas":{
        "Matemática": 99,
        "Geografia": 99,
        "Inglês": 98,
        "História": 98
    },
    "Gisleine": {
        "Matemática": 78,
        "Geografia": 73,
        "Inglês": 56,
        "História": 88
    }
}


wb = Workbook()
ws = wb.active
ws.title = "Notas"

headings = ['Nome'] + list(data['João'].keys())
ws.append(headings)

#preenche as cédulas com as notas e os nomes dos alunos
for pessoa in data:
    notas = list(data[pessoa].values())
    ws.append([pessoa] + notas)

#faz a média das notas
for col in range(2, len(data['João']) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

#estiliza as cédulas
for col in range(1,6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")


wb.save("NovoNotas.xlsx")